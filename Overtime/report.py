import pandas as pd
import os
import openpyxl
import re
import shutil

# Directory containing CSV files
csv_folder = os.path.join(os.getcwd(), 'report')
# Initialize a list to store paths of the filtered Excel files
file_paths = []

# Create a temporary directory for intermediate files
temp_directory = os.path.join(os.getcwd(), "temp")
if not os.path.exists(temp_directory):
    os.makedirs(temp_directory)
    print(f"Created temp directory: {temp_directory}")

# Iterate over all CSV files in the folder
for file in os.listdir(csv_folder):
    if file.endswith('.csv'):
        file_name = os.path.basename(file)
        base_name, _ = os.path.splitext(file_name)
        
        # Extract period from file name
        period_match = re.search(r'(\d{4}_\d{2}_\d{2})-(\d{4}_\d{2}_\d{2})', file)
        if period_match:
            start_date, end_date = period_match.groups()
            period_string = f'{start_date}-{end_date}'
        else:
            period_string = 'Unknown Period'
        
        file_path = os.path.join(csv_folder, file)
        print(f"Processing file: {file_path}")
        
        # Read CSV file
        try:
            df = pd.read_csv(file_path, encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file_path, encoding='ISO-8859-1')
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, encoding='windows-1252')

        # Determine sheet name based on 'Location' column
        if 'Location' in df.columns and df['Location'].str.contains('Carlsbad', case=False, na=False).any():
            sheet_name = 'CB'
        elif 'Location' in df.columns and df['Location'].str.contains('Oceanside', case=False, na=False).any():
            sheet_name = 'OS'
        elif 'Location' in df.columns and df['Location'].str.contains('La Jolla', case=False, na=False).any():
            sheet_name = 'LJ'
        elif 'Location' in df.columns and df['Location'].str.contains('Rancho Bernardo', case=False, na=False).any():
            sheet_name = 'RB'
        elif 'Location' in df.columns and df['Location'].str.contains('Del Mar', case=False, na=False).any():
            sheet_name = 'DM'
        elif 'Location' in df.columns and df['Location'].str.contains('Encinitas', case=False, na=False).any():
            sheet_name = 'EN'
        else:
            sheet_name = 'Unknown'
        output_file = os.path.join(temp_directory, f"{sheet_name}.xlsx")
        # Process DataFrame if it contains required columns
        if 'Employee' in df.columns and 'Overtime Hours' in df.columns and 'Job Title' in df.columns:
            # Remove rows with 'Job Title' containing 'Generic' or 'Driver'
            df = df[~df['Job Title'].str.contains('Generic|Driver', case=False, na=False)]
            # Process 'Half Day Server' and 'Server' job titles
            half_day_server_jobs = df[df['Job Title'].isin(['Half Day Server', 'Half Day Busser'])] 
            for _, half_day_row in half_day_server_jobs.iterrows():
                server_rows = df[(df['Job Title'].isin(['Server', 'Busser','Cashier'])) & (df['Employee'] == half_day_row['Employee'])]
                if not server_rows.empty:
                    idx = server_rows.index[0]
                    for column in df.columns:
                        if column not in ['Employee', 'Job Title','Hourly Rate','Location']:
                            df.at[idx, column] += half_day_row[column]
                    df = df[df.index != half_day_row.name]
            # Create 'Department' column based on 'Job Title'
            df['Department'] = df['Job Title'].apply(lambda x: 'Kitchen' if x in ['Drama', 'Chef', 'Dishwasher', 'Cook', 'Fryer'] else 'Front')
            # Convert 'Overtime Hours' to numeric, coerce errors to NaN
            df['Overtime Hours'] = pd.to_numeric(df['Overtime Hours'], errors='coerce')
            # Selecting only 'Department', 'Employee', and 'Overtime Hours' columns
            df = df[['Department', 'Employee', 'Overtime Hours']]
            
            # Define the sorting order for the 'Department' column
            df['Department'] = pd.Categorical(df['Department'], categories=['Front', 'Kitchen'], ordered=True)
            df = df.sort_values('Department')
            #Export the filtered DataFrame to an Excel file
            df.to_excel(output_file, index=False)

            df = pd.read_excel(output_file)           
            # Locate the indices where 'Front' and 'Kitchen' departments end
            last_front_index = df[df['Department'] == 'Front'].index[-1]
            last_kitchen_index = df[df['Department'] == 'Kitchen'].index[-1]

            # Calculate sum of 'Overtime Hours' for 'Front' and 'Kitchen'
            sum_front = df.loc[df['Department'] == 'Front', 'Overtime Hours'].sum()
            sum_kitchen = df.loc[df['Department'] == 'Kitchen', 'Overtime Hours'].sum()
            #new column
            df['Total'] = ''
            # Create sum rows and a blank row
            sum_front_row = pd.DataFrame([['', '','', sum_front]], columns=df.columns)
            blank_row = pd.DataFrame([['', '', '','']], columns=df.columns)
            sum_kitchen_row = pd.DataFrame([['', '','', sum_kitchen]], columns=df.columns)

            # Insert the sum row for 'Front' and the blank row
            df = pd.concat([df.iloc[:last_front_index + 1], sum_front_row, blank_row, df.iloc[last_front_index + 1:]]).reset_index(drop=True)

            # Append the sum row for 'Kitchen' at the end
            df = pd.concat([df, sum_kitchen_row]).reset_index(drop=True)
            df.to_excel(output_file, index=False)
            print(f'Filtered data exported to {output_file}')
            file_paths.append(output_file)

# Combine all filtered Excel files into one workbook
combined_file = os.path.join(os.getcwd(), 'overtime_filtered_data.xlsx')
with pd.ExcelWriter(combined_file, engine='openpyxl') as writer:
    for file_path in file_paths:
        df = pd.read_excel(file_path)
        
        # Use the file name without extension as the sheet name
        sheet_name = os.path.splitext(os.path.basename(file_path))[0]
        
        # Write DataFrame to a new sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Added {file_path} to {combined_file} as sheet {sheet_name}")

wb = openpyxl.load_workbook(combined_file)
pastel_pink = openpyxl.styles.PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
pastel_blue = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
# Iterate over all sheets in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    for row in ws.iter_rows(min_row=2):  # Start from the second row
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value, 2)
    total_col_index = None
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        if col[0].value == 'Total':
            total_col_index = col[0].column
            break

    if total_col_index is not None:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Assuming header is in the first row
            total_cell = row[total_col_index - 1]  # Adjust for 0-based index
            if total_cell.value is not None and total_cell.value != '':
                for cell in row:
                    cell.fill = pastel_pink

        total_sum = 0
        for cell in ws.iter_cols(min_col=total_col_index, max_col=total_col_index, min_row=2, max_row=ws.max_row):
            for c in cell:
                if c.value is not None and isinstance(c.value, (int, float)):
                    total_sum += c.value
                    print(f"Adding value {c.value} from cell {c.coordinate}")

        # Debug print to check the calculated sum
        print(f"Calculated sum of 'Total' column: {total_sum}")

        # Insert a new row at the end with the label 'Total' and the sum
        new_row_index = ws.max_row + 1
        ws.cell(row=new_row_index, column=total_col_index - 1, value='Total')
        ws.cell(row=new_row_index, column=total_col_index, value=total_sum)
    #Insert 'Period' row and blank row at the top
    ws.insert_rows(1, 2)  # Insert 2 rows at the top
    ws.cell(row=1, column=1, value='Period')
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
    ws.cell(row=1, column=2, value=period_string)

    bold_font = openpyxl.styles.Font(bold=True)
    ws.cell(row=1, column=1).font = bold_font

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if any(cell.value and 'total' in str(cell.value).lower() for cell in row):
                for cell in row:
                    cell.fill = pastel_blue

wb.save(combined_file)


print(f"All files combined into {combined_file}")
# Optionally clean up temporary files
shutil.rmtree(temp_directory)
print("Deleting temporary files")
