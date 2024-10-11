import os
import pandas as pd
import shutil
import openpyxl
# Directory containing CSV files
csv_folder = os.path.join(os.getcwd(), 'report')
# Initialize a list to store paths of the filtered Excel files
file_paths = []

# Create a temporary directory for intermediate files
temp_directory = os.path.join(os.getcwd(), "temp")
if not os.path.exists(temp_directory):
    os.makedirs(temp_directory)
    print(f"Created temp directory: {temp_directory}")

for file in os.listdir(csv_folder):
    if file.endswith('.csv'):
        file_path = os.path.join(csv_folder, file)
        print(f"Processing file: {file_path}")
        # Load the CSV file with error handling for encoding
        try:
            df = pd.read_csv(file_path, encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file_path, encoding='ISO-8859-1')
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, encoding='windows-1252')

        # Determine sheet name based on 'Location' column
        if 'Location' in df.columns and df['Location'].str.contains('Carlsbad', case=False, na=False).any():
            sheet_name = 'CB'
        elif 'Location' in df.columns and df['Location'].str.contains('Oceanside', case=False, na=False).any():
            sheet_name = 'OS'
        elif 'Location' in df.columns and df['Location'].str.contains('La Jolla', case=False, na=False).any():
            sheet_name = 'LJ'
        elif 'Location' in df.columns and df['Location'].str.contains('Rancho Bernardo', case=False, na=False).any():
            sheet_name = 'RB'
        elif 'Location' in df.columns and df['Location'].str.contains('Del Mar', case=False, na=False).any():
            sheet_name = 'DM'
        elif 'Location' in df.columns and df['Location'].str.contains('Encinitas', case=False, na=False).any():
            sheet_name = 'EN'
        else:
            sheet_name = 'Unknown'
        output_file = os.path.join(temp_directory, f"{sheet_name}.xlsx")
    # Filter rows where 'Job Title' contains 'Shift Manager' or 'Assistant Manager'
    if 'Job Title' in df.columns:
        filtered_df = df[df['Job Title'].str.contains('Shift Manager|Assistant Manager|General Manager|Owner', case=False, na=False)]
    else:
        filtered_df = pd.DataFrame()
    # Export the filtered data to a new CSV file
    filtered_df.to_excel(output_file, index=False)
    print(f'Filtered data exported to {output_file}')
    file_paths.append(output_file)

combined_file = os.path.join(os.getcwd(), 'Entry_time_filtered_data.xlsx')
with pd.ExcelWriter(combined_file, engine='openpyxl') as writer:
    for file_path in file_paths:
        df = pd.read_excel(file_path)
        
        # Use the file name without extension as the sheet name
        sheet_name = os.path.splitext(os.path.basename(file_path))[0]
        
        # Write DataFrame to a new sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Added {file_path} to {combined_file} as sheet {sheet_name}")

wb = openpyxl.load_workbook(combined_file)
pastel_pink = openpyxl.styles.PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
# Iterate over all sheets in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
wb.save(combined_file)
print(f"All files combined into {combined_file}")
# Optionally clean up temporary files
shutil.rmtree(temp_directory)
print("Deleting temporary files")