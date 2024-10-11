import pandas as pd
import os
import glob
import openpyxl

def reformat_employee_name(name):
    if isinstance(name, str):
        parts = name.split(', ')
        if len(parts) == 2:
            return f"{parts[1]} {parts[0]}"
    return name

def merge_time_and_tip_reports(time_report_file, tip_report_file, output_file):
    # Open the time report and tip report Excel files
    time_report = pd.ExcelFile(time_report_file)
    tip_report = pd.ExcelFile(tip_report_file)

    # Create a Pandas Excel writer object
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Loop through each sheet in the time report
        for sheet_name in time_report.sheet_names:
            if sheet_name in tip_report.sheet_names:
                # Read the time and tip sheets
                time_df = pd.read_excel(time_report, sheet_name=sheet_name)
                tip_df = pd.read_excel(tip_report, sheet_name=sheet_name)
                # Strip leading/trailing whitespace from 'Employee' columns
                time_df['Employee'] = time_df['Employee'].str.strip()
                tip_df['Employee'] = tip_df['Employee'].str.strip()

                # Reformat employee names in time_df to match tip_df
                time_df['Employee'] = time_df['Employee'].apply(reformat_employee_name)
            else:
                print(f"Sheet {sheet_name} not found in tip report.")
                continue

                # Ensure 'Employee' column is the key for merging
            if 'Employee' in time_df.columns and 'Employee' in tip_df.columns:
                # Merge the dataframes on the 'Employee' column
                merged_df = pd.merge(time_df, tip_df[['Employee','Job Title','Non-cash tips after pooling','Cash tips before pooling','Total tips after pooling' ]], on=['Employee','Job Title'], how='left')

                job_titles_for_deposit = ['Cook', 'Fryer', 'Dramma', 'Dishwasher']
                merged_df['CC tips (Deposit)'] = merged_df.apply(
                    lambda row: row['Non-cash tips after pooling'] if row['Job Title'] in job_titles_for_deposit else row['Non-cash tips after pooling'],
                    axis=1
                )
            else:
                print(f"Sheet {sheet_name} does not have 'Employee' column in both time and tip reports.")
                continue
            #rename cash tips column           
            merged_df.rename(columns={'Cash tips before pooling': 'Cash tips'}, inplace=True)
            # Move the 'CC tips' column to the end
            cash_tips = merged_df.pop('Cash tips')
            merged_df['Cash tips'] = cash_tips
            # Move the 'total tip' column to the end
            total_tips = merged_df.pop('Total tips after pooling')
            merged_df['Total tips after pooling'] = total_tips
            #drop unneeded columns    
            merged_df.drop(['Non-cash tips after pooling','Net Sales','Declared Tips','Non-Cash Tips','Total Tips','Tips Withheld','Total Gratuity','Employee ID','Job Code','Location Code'], axis=1, inplace=True)
            #move location column
            columns = ['Location'] + [col for col in merged_df.columns if col != 'Location']
            merged_df = merged_df[columns]
            # Filter out all blank rows
            merged_df.dropna(how='all', inplace=True)
            # Add 'Total Per Employee' column
            numeric_columns = ['Regular Pay', 'Overtime Pay', 'Total tips after pooling']  # Add other relevant numeric columns if needed
            merged_df['Total Per Employee'] = merged_df[numeric_columns].sum(axis=1)
            # Add sum row
            sum_row = merged_df.sum(numeric_only=True)
            sum_row['Job Title'] = 'Total'  # Set a label for the sum row
            # Append the sum row to the DataFrame
            merged_df = pd.concat([merged_df, sum_row.to_frame().T], ignore_index=True)

            # Write the merged dataframe to the output file
            merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Merged sheet {sheet_name} and written to output file.")
    
                
# Define the file paths
file_path=os.path.join(os.getcwd(),'output')
# Define the pattern to search for files containing 'PayrollExport'
pattern = os.path.join(file_path, '*PayrollExport*.xlsx')
matching_files = glob.glob(pattern)

# Check if any matching files were found
if matching_files:
    time_report_file = matching_files[0]  # Select the first matching file
    print(f"Found time report file: {time_report_file}")
    # Extract the base name and construct the new output file name
    base_name = os.path.basename(time_report_file)
    name, _ = os.path.splitext(base_name)
    output_file = os.path.join(file_path, f"{name}_merged.xlsx")
    print(f"Output file: {output_file}")
else:
    raise FileNotFoundError("No file containing 'PayrollExport' was found.")
tip_report_file = os.path.join(file_path,'tip_filtered_data.xlsx')

merge_time_and_tip_reports(time_report_file, tip_report_file, output_file)
#remove files
os.remove(time_report_file)
os.remove(tip_report_file)
wb = openpyxl.load_workbook(output_file)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    #apply formular
    max_row = ws.max_row
    for row in range(2, max_row + 1):  # Assuming the first row is the header
        title = ws[f'C{row}'].value  # Check if column C cell is empty
        if title is not None:  # If the cell is not empty
            regular_pay = f'=ROUND(D{row}*F{row},2)'
            ws[f'G{row}'] = regular_pay  
            overtime_pay = f'=ROUND(E{row}*F{row}*1.5,2)'
            ws[f'H{row}'] = overtime_pay 
            total_pay = f'=ROUND(G{row}+H{row},2)'
            ws[f'I{row}'] = total_pay 
            total_tip=f'=ROUND(J{row}+K{row},2)'
            ws[f'L{row}']=total_tip
            total=f'=ROUND(I{row}+L{row},2)'
            ws[f'M{row}']=total
        else:
            continue
    max_col=ws.max_column
    for col in range(4, max_col + 1):
    # Construct column letter (D, E, F, etc.)
        col_letter = ws.cell(row=1, column=col).column_letter       
        # Define the cell in row 31 for the sum
        sum_cell = ws[f'{col_letter}{max_row}'] 
        # Define the range to sum (excluding row 1)
        sum_range = f'{col_letter}2:{col_letter}{max_row-1}'
        
        # Write the SUM formula to cell in row 31
        sum_cell.value = f'=ROUND(SUM({sum_range}),2)'
        

wb.save(output_file)